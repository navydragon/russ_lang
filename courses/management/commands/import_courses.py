import os
from pathlib import Path
from django.core.management.base import BaseCommand
from django.db import connection, transaction
from openpyxl import load_workbook
from courses.models import Course, Lesson, Task


class Command(BaseCommand):
    help = 'Очищает таблицы курсов, уроков и заданий, сбрасывает счетчики и импортирует данные из Excel файла'

    def add_arguments(self, parser):
        parser.add_argument(
            'file_path',
            type=str,
            help='Путь к Excel файлу import_courses.xlsx (по умолчанию ищется в папке с командой или в корне проекта)',
            nargs='?',
            default=None
        )

    def handle(self, *args, **options):
        file_path = options['file_path']
        
        # Если путь не указан, ищем файл в нескольких местах
        if not file_path:
            # Сначала проверяем в папке с командой
            command_dir = Path(__file__).parent
            default_path = command_dir / 'import_courses.xlsx'
            
            if default_path.exists():
                file_path = str(default_path)
            else:
                # Потом в корне проекта (где manage.py)
                base_dir = Path(__file__).parent.parent.parent.parent
                default_path = base_dir / 'import_courses.xlsx'
                
                if default_path.exists():
                    file_path = str(default_path)
                else:
                    file_path = 'import_courses.xlsx'  # Попытка найти в текущей директории
        
        # Проверяем существование файла
        if not os.path.exists(file_path):
            self.stdout.write(
                self.style.ERROR(
                    f'Файл {file_path} не найден!\n'
                    f'Поместите файл import_courses.xlsx в одну из следующих директорий:\n'
                    f'  - {Path(__file__).parent} (папка с командой)\n'
                    f'  - {Path(__file__).parent.parent.parent.parent} (корень проекта)\n'
                    f'Или укажите полный путь к файлу при вызове команды.'
                )
            )
            return

        self.stdout.write(self.style.SUCCESS(f'Начинаю импорт из файла: {file_path}'))

        try:
            with transaction.atomic():
                # 1. Очищаем таблицы (в правильном порядке из-за внешних ключей)
                self.stdout.write('Очистка таблиц...')
                Task.objects.all().delete()
                Lesson.objects.all().delete()
                Course.objects.all().delete()

                # 2. Сбрасываем счетчики автоинкремента
                db_engine = connection.vendor
                with connection.cursor() as cursor:
                    if db_engine == 'sqlite':
                        # SQLite
                        cursor.execute("DELETE FROM sqlite_sequence WHERE name IN ('courses_task', 'courses_lesson', 'courses_course')")
                    elif db_engine == 'postgresql':
                        # PostgreSQL
                        cursor.execute("ALTER SEQUENCE courses_course_id_seq RESTART WITH 1")
                        cursor.execute("ALTER SEQUENCE courses_lesson_id_seq RESTART WITH 1")
                        cursor.execute("ALTER SEQUENCE courses_task_id_seq RESTART WITH 1")
                    elif db_engine == 'mysql':
                        # MySQL
                        cursor.execute("ALTER TABLE courses_course AUTO_INCREMENT = 1")
                        cursor.execute("ALTER TABLE courses_lesson AUTO_INCREMENT = 1")
                        cursor.execute("ALTER TABLE courses_task AUTO_INCREMENT = 1")

                self.stdout.write(self.style.SUCCESS('Таблицы очищены, счетчики сброшены'))

                # 3. Загружаем Excel файл
                workbook = load_workbook(file_path, data_only=True)

                # 4. Импортируем курсы (лист "courses")
                self.stdout.write('Импорт курсов...')
                if 'courses' not in workbook.sheetnames:
                    raise ValueError('Лист "courses" не найден в Excel файле!')

                courses_sheet = workbook['courses']
                course_name = None

                # Читаем название курса из второй строки (первая - заголовок "name")
                for row in courses_sheet.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # Если есть значение в первой колонке
                        course_name = str(row[0]).strip()
                        break

                if not course_name:
                    raise ValueError('Не найдено название курса в листе "courses"!')

                course = Course.objects.create(name=course_name)
                self.stdout.write(self.style.SUCCESS(f'  ✓ Создан курс: {course_name}'))

                # 5. Импортируем уроки (лист "lessons")
                self.stdout.write('Импорт уроков...')
                if 'lessons' not in workbook.sheetnames:
                    raise ValueError('Лист "lessons" не найден в Excel файле!')

                lessons_sheet = workbook['lessons']
                lesson_mapping = {}  # Словарь для связи имени урока с объектом

                # Читаем уроки (первая строка - заголовки)
                for row in lessons_sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1] and row[2]:  # name, course, position
                        lesson_name = str(row[0]).strip()
                        course_name_check = str(row[1]).strip()
                        position = int(row[2])

                        # Проверяем, что урок относится к правильному курсу
                        if course_name_check != course_name:
                            self.stdout.write(
                                self.style.WARNING(
                                    f'  ⚠ Урок "{lesson_name}" относится к курсу "{course_name_check}", '
                                    f'но ожидается "{course_name}". Пропускаю...'
                                )
                            )
                            continue

                        lesson = Lesson.objects.create(
                            name=lesson_name,
                            course=course,
                            position=position
                        )
                        lesson_mapping[lesson_name] = lesson
                        self.stdout.write(self.style.SUCCESS(f'  ✓ Создан урок: {lesson_name} (позиция {position})'))

                # 6. Импортируем задания (лист "tasks")
                self.stdout.write('Импорт заданий...')
                if 'tasks' not in workbook.sheetnames:
                    raise ValueError('Лист "tasks" не найден в Excel файле!')

                tasks_sheet = workbook['tasks']
                tasks_count = 0

                # Читаем задания (первая строка - заголовки)
                for row in tasks_sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[2] and row[3]:  # name, lesson, position (code может быть пустым)
                        task_name = str(row[0]).strip()
                        task_code = str(row[1]).strip() if row[1] else ''
                        lesson_name = str(row[2]).strip()
                        position = int(row[3])

                        # Находим урок по имени
                        lesson = lesson_mapping.get(lesson_name)
                        if not lesson:
                            self.stdout.write(
                                self.style.WARNING(
                                    f'  ⚠ Урок "{lesson_name}" не найден для задания "{task_name}". Пропускаю...'
                                )
                            )
                            continue

                        Task.objects.create(
                            name=task_name,
                            code=task_code,
                            lesson=lesson,
                            position=position
                        )
                        tasks_count += 1
                        self.stdout.write(
                            self.style.SUCCESS(
                                f'  ✓ Создано задание: {task_name} (код: {task_code}, позиция {position})'
                            )
                        )

                self.stdout.write(self.style.SUCCESS(f'\nИмпорт завершен успешно!'))
                self.stdout.write(
                    self.style.SUCCESS(
                        f'Импортировано: 1 курс, {len(lesson_mapping)} уроков, {tasks_count} заданий'
                    )
                )

        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'Ошибка при импорте: {str(e)}')
            )
            raise

