import os
from collections import defaultdict
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction
from django.db.models import Max
from openpyxl import load_workbook

from courses.models import Lesson, Task


class Command(BaseCommand):
    help = 'Импортирует задания из Excel (код, название, id урока)'

    def add_arguments(self, parser):
        parser.add_argument(
            'file_path',
            type=str,
            nargs='?',
            default=None,
            help='Путь к Excel файлу (по умолчанию courses/Книга русский.xlsx)',
        )
        parser.add_argument(
            '--keep-existing',
            action='store_true',
            help='Не удалять существующие задания у импортируемых уроков',
        )

    def handle(self, *args, **options):
        file_path = options['file_path']

        if not file_path:
            base_dir = Path(__file__).parent.parent.parent.parent
            default_path = base_dir / 'Книга русский.xlsx'
            file_path = str(default_path) if default_path.exists() else 'courses/Книга русский.xlsx'

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Файл {file_path} не найден!'))
            return

        self.stdout.write(self.style.SUCCESS(f'Импорт заданий из: {file_path}'))

        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active

        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            code = str(row[0]).strip()
            name = str(row[1]).strip() if row[1] else ''
            if not name:
                num = code.rsplit('-', 1)[-1]
                name = f'Задание {num}'
            lesson_id = int(row[2])
            rows.append({'code': code, 'name': name, 'lesson_id': lesson_id})

        if not rows:
            self.stdout.write(self.style.WARNING('В файле нет строк с заданиями'))
            return

        lesson_ids = {row['lesson_id'] for row in rows}
        lessons = {lesson.id: lesson for lesson in Lesson.objects.filter(id__in=lesson_ids)}
        missing_lessons = lesson_ids - lessons.keys()
        if missing_lessons:
            self.stdout.write(
                self.style.ERROR(
                    f'Уроки с id {sorted(missing_lessons)} не найдены в базе. Импорт прерван.'
                )
            )
            return

        position_counters = defaultdict(int)
        existing_codes = set()

        if options['keep_existing']:
            for item in Task.objects.filter(
                lesson_id__in=lesson_ids
            ).values('lesson_id').annotate(max_pos=Max('position')):
                position_counters[item['lesson_id']] = item['max_pos'] or 0
            existing_codes = set(
                Task.objects.filter(lesson_id__in=lesson_ids).values_list(
                    'lesson_id', 'code'
                )
            )

        try:
            with transaction.atomic():
                if not options['keep_existing']:
                    deleted, _ = Task.objects.filter(
                        lesson_id__in=lesson_ids
                    ).delete()
                    self.stdout.write(f'Удалено заданий: {deleted}')

                created = 0
                skipped = 0
                for row in rows:
                    lesson = lessons[row['lesson_id']]
                    if options['keep_existing'] and (lesson.id, row['code']) in existing_codes:
                        skipped += 1
                        self.stdout.write(
                            self.style.WARNING(
                                f'  ~ пропуск {row["code"]} | урок {lesson.id} — уже существует'
                            )
                        )
                        continue

                    position_counters[lesson.id] += 1
                    position = position_counters[lesson.id]

                    Task.objects.create(
                        name=row['name'],
                        code=row['code'],
                        lesson=lesson,
                        position=position,
                    )
                    existing_codes.add((lesson.id, row['code']))
                    created += 1
                    self.stdout.write(
                        self.style.SUCCESS(
                            f'  + {row["code"]} | {row["name"]} '
                            f'| урок {lesson.id} | поз. {position}'
                        )
                    )

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Ошибка при импорте: {e}'))
            raise

        self.stdout.write(
            self.style.SUCCESS(
                f'\nГотово: создано {created}, пропущено {skipped} из {len(rows)} строк'
            )
        )
