import os
import re
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand
from django.db import connection, transaction
from django.utils.crypto import get_random_string
from openpyxl import load_workbook

from courses.models import StudentTask, TaskAttempt
from students.models import Group, Student

User = get_user_model()

DEFAULT_FILE = 'config/Русский.xlsx'

COLUMNS = {
    'group': 'Группа',
    'last_name': 'Фамилия',
    'first_name': 'Имя',
    'middle_name': 'Отчество',
    'email': 'Почта',
    'curator': 'Куратор',
    'curator_email': 'Почта Куратора',
    'login': 'login',
    'code': 'code',
    'password': 'password',
}


def normalize(value):
    if value is None:
        return ''
    return str(value).strip()


def parse_curator_name(raw_name):
    raw_name = ' '.join(normalize(raw_name).split())
    if not raw_name:
        return '', '', None

    parts = raw_name.split(' ', 1)
    last_name = parts[0]
    if len(parts) == 1:
        return last_name, '', None

    letters = [part for part in re.split(r'[.\s]+', parts[1]) if part]
    first_name = f'{letters[0]}.' if letters else ''
    middle_name = f'{letters[1]}.' if len(letters) > 1 else None
    return last_name, first_name, middle_name


def group_code_from_name(group_name):
    match = re.search(r'(\d+)', group_name)
    if match:
        return f'gr-{match.group(1)}'
    slug = re.sub(r'[^\w-]+', '-', group_name.lower()).strip('-')
    return slug or 'group'


def username_from_email(email, existing_usernames):
    base = re.sub(r'[^\w.@+-]', '_', email.split('@')[0].lower())
    username = base
    suffix = 1
    while username in existing_usernames:
        username = f'{base}{suffix}'
        suffix += 1
    existing_usernames.add(username)
    return username


class Command(BaseCommand):
    help = (
        'Импортирует кураторов (пользователей), группы и студентов из Excel '
        'с корректными связями Group.curators и Student.group'
    )

    def add_arguments(self, parser):
        parser.add_argument(
            'file_path',
            nargs='?',
            default=None,
            help=f'Путь к Excel-файлу (по умолчанию: {DEFAULT_FILE})',
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Очистить студентов, группы и связанные попытки перед импортом',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Показать, что будет создано, без записи в БД',
        )
        parser.add_argument(
            '--curator-password',
            default=None,
            help='Пароль для новых кураторов (если не указан — генерируется случайный)',
        )

    def resolve_file_path(self, file_path):
        if file_path:
            return file_path

        base_dir = Path(__file__).resolve().parents[3]
        default_path = base_dir / DEFAULT_FILE
        if default_path.exists():
            return str(default_path)

        return DEFAULT_FILE

    def clear_data(self):
        self.stdout.write('Очистка данных студентов и групп...')
        TaskAttempt.objects.all().delete()
        StudentTask.objects.all().delete()
        Student.objects.all().delete()
        Group.objects.all().delete()

        db_engine = connection.vendor
        with connection.cursor() as cursor:
            if db_engine == 'sqlite':
                cursor.execute(
                    "DELETE FROM sqlite_sequence WHERE name IN "
                    "('students_student', 'students_group')"
                )
            elif db_engine == 'postgresql':
                cursor.execute('ALTER SEQUENCE students_group_id_seq RESTART WITH 1')
                cursor.execute('ALTER SEQUENCE students_student_id_seq RESTART WITH 1')
            elif db_engine == 'mysql':
                cursor.execute('ALTER TABLE students_group AUTO_INCREMENT = 1')
                cursor.execute('ALTER TABLE students_student AUTO_INCREMENT = 1')

        self.stdout.write(self.style.SUCCESS('Данные очищены'))

    def read_rows(self, file_path):
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active

        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        header_map = {normalize(value): index for index, value in enumerate(header_row)}

        missing = [title for title in COLUMNS.values() if title not in header_map]
        if missing:
            raise ValueError(f'В файле отсутствуют колонки: {", ".join(missing)}')

        rows = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            record = {
                key: normalize(row[header_map[title]]) if row[header_map[title]] is not None else ''
                for key, title in COLUMNS.items()
            }
            if not any(record.values()):
                continue
            record['row_number'] = row_number
            rows.append(record)

        return rows

    def handle(self, *args, **options):
        file_path = self.resolve_file_path(options['file_path'])

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Файл не найден: {file_path}'))
            return

        self.stdout.write(self.style.SUCCESS(f'Импорт из файла: {file_path}'))

        try:
            rows = self.read_rows(file_path)
        except Exception as exc:
            self.stdout.write(self.style.ERROR(f'Ошибка чтения файла: {exc}'))
            return

        if not rows:
            self.stdout.write(self.style.WARNING('В файле нет данных для импорта'))
            return

        curators_data = {}
        groups_data = {}
        students_data = []
        seen_codes = set()

        for row in rows:
            if not row['code']:
                self.stdout.write(
                    self.style.WARNING(
                        f'Строка {row["row_number"]}: пропущена — нет code'
                    )
                )
                continue

            if row['code'] in seen_codes:
                self.stdout.write(
                    self.style.WARNING(
                        f'Строка {row["row_number"]}: дубликат code "{row["code"]}", пропуск'
                    )
                )
                continue
            seen_codes.add(row['code'])

            curator_email = row['curator_email'].lower()
            if not curator_email:
                self.stdout.write(
                    self.style.WARNING(
                        f'Строка {row["row_number"]}: пропущена — нет почты куратора'
                    )
                )
                continue

            if not row['group']:
                self.stdout.write(
                    self.style.WARNING(
                        f'Строка {row["row_number"]}: пропущена — нет группы'
                    )
                )
                continue

            if curator_email not in curators_data:
                last_name, first_name, middle_name = parse_curator_name(row['curator'])
                curators_data[curator_email] = {
                    'email': curator_email,
                    'last_name': last_name,
                    'first_name': first_name,
                    'middle_name': middle_name,
                    'display_name': row['curator'],
                }

            group_name = row['group']
            if group_name not in groups_data:
                groups_data[group_name] = {
                    'name': group_name,
                    'code': group_code_from_name(group_name),
                    'curator_emails': set(),
                }
            groups_data[group_name]['curator_emails'].add(curator_email)

            students_data.append({
                'last_name': row['last_name'],
                'first_name': row['first_name'],
                'middle_name': row['middle_name'] or None,
                'code': row['code'],
                'group_name': group_name,
                'row_number': row['row_number'],
            })

        self.stdout.write(
            f'Найдено: {len(curators_data)} кураторов, '
            f'{len(groups_data)} групп, {len(students_data)} студентов'
        )

        if options['dry_run']:
            for email, curator in curators_data.items():
                self.stdout.write(f'  куратор: {curator["display_name"]} <{email}>')
            for group in groups_data.values():
                curator_list = ', '.join(sorted(group['curator_emails']))
                self.stdout.write(
                    f'  группа: {group["name"]} ({group["code"]}) -> {curator_list}'
                )
            self.stdout.write(self.style.SUCCESS('Dry-run завершён, изменения не сохранены'))
            return

        created_passwords = []

        try:
            with transaction.atomic():
                if options['clear']:
                    self.clear_data()

                existing_usernames = set(User.objects.values_list('username', flat=True))
                curators_by_email = {}

                for email, curator in curators_data.items():
                    user = User.objects.filter(email__iexact=email).first()
                    if user:
                        user.last_name = curator['last_name'] or user.last_name
                        user.first_name = curator['first_name'] or user.first_name
                        user.middle_name = curator['middle_name'] or user.middle_name
                        user.is_tutor = True
                        user.send_emails = True
                        user.save()
                        self.stdout.write(f'  ~ обновлён куратор: {user}')
                    else:
                        username = username_from_email(email, existing_usernames)
                        password = options['curator_password'] or get_random_string(12)
                        user = User.objects.create_user(
                            username=username,
                            email=email,
                            password=password,
                            last_name=curator['last_name'],
                            first_name=curator['first_name'],
                            middle_name=curator['middle_name'],
                            is_tutor=True,
                            send_emails=True,
                        )
                        created_passwords.append((user, password))
                        self.stdout.write(f'  + создан куратор: {user} ({email})')

                    curators_by_email[email] = user

                groups_by_name = {}
                used_codes = set(Group.objects.values_list('code', flat=True))

                for group_name, group_data in groups_data.items():
                    code = group_data['code']
                    if code in used_codes:
                        suffix = 2
                        while f'{code}-{suffix}' in used_codes:
                            suffix += 1
                        code = f'{code}-{suffix}'

                    group = Group.objects.create(
                        name=group_data['name'],
                        code=code,
                    )
                    group.curators.set([
                        curators_by_email[email] for email in group_data['curator_emails']
                    ])
                    used_codes.add(code)
                    groups_by_name[group_name] = group
                    self.stdout.write(
                        f'  + создана группа: {group.name} ({group.code})'
                    )

                for student_data in students_data:
                    group = groups_by_name[student_data['group_name']]
                    Student.objects.create(
                        last_name=student_data['last_name'],
                        first_name=student_data['first_name'],
                        middle_name=student_data['middle_name'],
                        code=student_data['code'],
                        group=group,
                    )

                self.stdout.write(
                    self.style.SUCCESS(
                        f'\nИмпорт завершён: {len(curators_by_email)} кураторов, '
                        f'{len(groups_by_name)} групп, {len(students_data)} студентов'
                    )
                )

                if created_passwords:
                    self.stdout.write('\nПароли новых кураторов:')
                    for user, password in created_passwords:
                        self.stdout.write(f'  {user.username} ({user.email}): {password}')

        except Exception as exc:
            self.stdout.write(self.style.ERROR(f'Ошибка импорта: {exc}'))
            raise
